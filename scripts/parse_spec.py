#!/usr/bin/env python3
"""Parse SDTM spec xlsx: extract filtered variables or write algorithm descriptions back."""

import argparse
import json
import sys
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

REQUIRED_COLUMNS = [
    "DATASET",
    "TAGGED_FORM",
    "VARIABLE",
    "ALGORITHM_STATUS",
    "TRANSFORMATION_TYPE",
    "ENGLISH_ALGORITHM_DESCRIPTION",
]

OPTIONAL_COLUMNS = [
    "POST_MACRO_CUSTOM_CHANGES",
    "POST_MACR_CUSTOM_CHANGES",
]

FILTER_STATUS = {"CHGREQ", "CHGOPT"}
FILTER_TYPE = {"CUSTOM", "ZD_DM", "LBSTRESC_CLRM", "VISITNUM_CLRM", "DIRECT_CONDITIONAL", "DIRECT"}


def fuzzy_match_column(header, target):
    """Match column names allowing for common typos."""
    if header is None:
        return False
    h = str(header).strip().upper()
    t = target.upper()
    if h == t:
        return True
    # Handle POST_MACR vs POST_MACRO typo
    if "POST_MAC" in t and "POST_MAC" in h:
        return True
    return False


def find_sheet(wb, sas_name):
    """Find matching sheet by SAS filename stem (case-insensitive)."""
    target = sas_name.upper()
    for name in wb.sheetnames:
        if name.upper() == target:
            return name
    # Try partial match
    for name in wb.sheetnames:
        if target in name.upper() or name.upper() in target:
            return name
    return None


def detect_columns(header_row):
    """Detect column positions from header row. Returns dict of column_name -> index."""
    col_map = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        cell_str = str(cell).strip().upper()
        for col in REQUIRED_COLUMNS:
            if fuzzy_match_column(cell_str, col):
                col_map[col] = idx
                break
        else:
            for col in OPTIONAL_COLUMNS:
                if fuzzy_match_column(cell_str, col):
                    col_map["POST_MACRO_CUSTOM_CHANGES"] = idx
                    break
    return col_map


def extract(spec_path, sas_name):
    """Extract filtered variables from spec. Outputs JSON to stdout."""
    wb = openpyxl.load_workbook(spec_path, read_only=True, data_only=True)

    sheet_name = find_sheet(wb, sas_name)
    if not sheet_name:
        print(json.dumps({"error": f"No sheet matching '{sas_name}' found. Available: {wb.sheetnames}"}))
        sys.exit(1)

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(json.dumps({"error": "Sheet is empty"}))
        sys.exit(1)

    header = rows[0]
    col_map = detect_columns(header)

    missing = [c for c in ["VARIABLE", "ALGORITHM_STATUS", "TRANSFORMATION_TYPE"] if c not in col_map]
    if missing:
        print(json.dumps({"error": f"Missing required columns: {missing}. Found: {list(col_map.keys())}"}))
        sys.exit(1)

    results = []
    for row_idx, row in enumerate(rows[1:], start=2):
        status = str(row[col_map["ALGORITHM_STATUS"]] or "").strip().upper()
        trans_type = str(row[col_map["TRANSFORMATION_TYPE"]] or "").strip().upper()

        if status in FILTER_STATUS and trans_type in FILTER_TYPE:
            record = {"row_index": row_idx}
            for col_name, col_idx in col_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                record[col_name] = str(val) if val is not None else ""
            results.append(record)

    output = {
        "sheet_name": sheet_name,
        "sas_name": sas_name,
        "total_rows": len(rows) - 1,
        "filtered_count": len(results),
        "columns_detected": list(col_map.keys()),
        "variables": results,
    }
    print(json.dumps(output, indent=2))
    wb.close()


def write_output(spec_path, sas_name, output_path, algorithms_json, spec_driven=False):
    """Write output xlsx with algorithm descriptions populated."""
    wb = openpyxl.load_workbook(spec_path, read_only=True, data_only=True)

    sheet_name = find_sheet(wb, sas_name)
    if not sheet_name:
        print(f"ERROR: No sheet matching '{sas_name}'", file=sys.stderr)
        sys.exit(1)

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    col_map = detect_columns(header)
    wb.close()

    with open(algorithms_json, "r") as f:
        algorithms = json.load(f)

    # Build lookup for spec-driven mode: key = (DATASET, TAGGED_FORM, VARIABLE)
    algo_lookup = {}
    if spec_driven and "variables" in algorithms:
        for entry in algorithms["variables"]:
            key = (
                str(entry.get("DATASET", "")).strip().upper(),
                str(entry.get("TAGGED_FORM", "")).strip().upper(),
                str(entry.get("VARIABLE", "")).strip().upper(),
            )
            algo_lookup[key] = entry.get("ENGLISH_ALGORITHM_DESCRIPTION", "")
    elif not spec_driven:
        # Legacy mode: simple variable name -> description dict
        if isinstance(algorithms, dict) and "variables" not in algorithms:
            algo_lookup = algorithms
        elif "variables" in algorithms:
            for entry in algorithms["variables"]:
                var = str(entry.get("VARIABLE", "")).strip()
                algo_lookup[var] = entry.get("ENGLISH_ALGORITHM_DESCRIPTION", "")

    # Create output workbook
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = sheet_name

    # Write header
    header_list = list(header)
    out_ws.append([str(h) if h else "" for h in header_list])

    # Write rows with algorithms populated
    algo_col = col_map.get("ENGLISH_ALGORITHM_DESCRIPTION")
    var_col = col_map.get("VARIABLE")
    ds_col = col_map.get("DATASET")
    tf_col = col_map.get("TAGGED_FORM")

    populated_count = 0
    rows_written = 0

    for row_idx, row in enumerate(rows[1:], start=2):
        if spec_driven:
            # Spec-driven mode: include row if it has a matching algorithm
            dataset = str(row[ds_col] or "").strip().upper() if ds_col is not None else ""
            tagged_form = str(row[tf_col] or "").strip().upper() if tf_col is not None else ""
            variable = str(row[var_col] or "").strip().upper() if var_col is not None else ""
            key = (dataset, tagged_form, variable)

            if key in algo_lookup:
                row_list = list(row)
                while len(row_list) < len(header_list):
                    row_list.append(None)
                if algo_col is not None:
                    row_list[algo_col] = algo_lookup[key]
                    populated_count += 1
                out_ws.append([str(v) if v is not None else "" for v in row_list])
                rows_written += 1
        else:
            # Standard filter mode
            status = str(row[col_map["ALGORITHM_STATUS"]] or "").strip().upper()
            trans_type = str(row[col_map["TRANSFORMATION_TYPE"]] or "").strip().upper()

            if status in FILTER_STATUS and trans_type in FILTER_TYPE:
                row_list = list(row)
                while len(row_list) < len(header_list):
                    row_list.append(None)

                variable = str(row_list[var_col] or "").strip()
                if variable in algo_lookup and algo_col is not None:
                    row_list[algo_col] = algo_lookup[variable]
                    populated_count += 1

                out_ws.append([str(v) if v is not None else "" for v in row_list])
                rows_written += 1

    # Auto-adjust column widths
    for col in out_ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        out_ws.column_dimensions[col_letter].width = max_len + 2

    out_wb.save(output_path)
    print(f"Output written to: {output_path}")
    print(f"Rows written: {rows_written}")
    print(f"Variables populated: {populated_count}")


def main():
    parser = argparse.ArgumentParser(description="SDTM Spec Parser for Algorithm Generation")
    parser.add_argument("--mode", choices=["extract", "write"], required=True)
    parser.add_argument("--spec", required=True, help="Path to spec xlsx file")
    parser.add_argument("--sas-name", required=True, help="SAS filename stem (e.g., KIAC_SE)")
    parser.add_argument("--output", help="Output xlsx path (write mode)")
    parser.add_argument("--algorithms", help="Path to JSON file with algorithms (write mode)")
    parser.add_argument("--spec-driven", action="store_true",
                        help="Use spec-driven matching (DATASET+TAGGED_FORM+VARIABLE) instead of status/type filters. For LBXL/ZDDM domains.")

    args = parser.parse_args()

    if not os.path.exists(args.spec):
        print(f"ERROR: Spec file not found: {args.spec}", file=sys.stderr)
        sys.exit(1)

    if args.mode == "extract":
        extract(args.spec, args.sas_name)
    elif args.mode == "write":
        if not args.output or not args.algorithms:
            print("ERROR: --output and --algorithms required for write mode", file=sys.stderr)
            sys.exit(1)
        write_output(args.spec, args.sas_name, args.output, args.algorithms, spec_driven=args.spec_driven)


if __name__ == "__main__":
    main()
